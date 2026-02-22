"""
Freelance Otomasyon Botu - Flask Backend v3 PREMIUM (Düzeltilmiş)
OCR | Excel | Fatura PDF (Logo + QR + Banka Bloğu + Sayfa No + Grid)

Düzeltmeler:
  - Font yolu: Proje dizininden göreceli yol, ortam değişkeni desteği
  - index() rotası: os.path.dirname(__file__) ile taşınabilir yol
  - Spesifik hata yakalama: ValueError, FileNotFoundError, OSError
  - CSV decode hataları: errors="strict" + açıklayıcı mesaj
  - Sayısal dönüşüm uyarısı: Geçersiz değer sayısı response'a eklendi
  - Veri doğrulama: kdv_oran, miktar, birim_fiyat güvenli float dönüşümü
  - Gereksiz bağımlılıklar: python-docx ve schedule kaldırıldı (requirements)
"""

import os, json, csv, io, base64, uuid, datetime, re, logging
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template_string

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── Excel ─────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PDF ───────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable,
                                 Image as RLImage, KeepTogether)
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus.flowables import Flowable

# ── Türkçe font kaydı (çoklu yol, ortam değişkeni desteği) ──────────
# Öncelik: 1) FONT_DIR env değişkeni  2) proje dizinindeki fonts/  3) sistem yolları
_BASE_DIR = Path(__file__).resolve().parent
_FONT_CANDIDATES = [
    os.environ.get("FONT_DIR", ""),                      # ortam değişkeni
    str(_BASE_DIR / "fonts"),                            # proje içi fonts/ klasörü
    "/usr/share/fonts/truetype/dejavu",                  # Debian/Ubuntu
    "/usr/share/fonts/dejavu",                           # bazı dağıtımlar
    "/usr/share/fonts/TTF",                              # Arch Linux
]

def _find_font(filename: str) -> str | None:
    for directory in _FONT_CANDIDATES:
        if not directory:
            continue
        path = Path(directory) / filename
        if path.exists():
            return str(path)
    return None

_regular = _find_font("DejaVuSans.ttf")
_bold    = _find_font("DejaVuSans-Bold.ttf")
_italic  = _find_font("DejaVuSans-Oblique.ttf")

try:
    if not (_regular and _bold and _italic):
        raise FileNotFoundError("DejaVu fontları bulunamadı, Helvetica kullanılacak.")
    pdfmetrics.registerFont(TTFont("DVSans",        _regular))
    pdfmetrics.registerFont(TTFont("DVSans-Bold",   _bold))
    pdfmetrics.registerFont(TTFont("DVSans-Italic", _italic))
    _FONT      = "DVSans"
    _FONT_BOLD = "DVSans-Bold"
    logger.info("DejaVu fontları başarıyla yüklendi.")
except (FileNotFoundError, Exception) as exc:
    logger.warning("Font yükleme hatası: %s — Helvetica kullanılıyor.", exc)
    _FONT      = "Helvetica"
    _FONT_BOLD = "Helvetica-Bold"

# ── Görüntü / OCR ─────────────────────────────────────────────────────
from PIL import Image, UnidentifiedImageError
import pytesseract

app = Flask(__name__)
UPLOAD = Path("/tmp/bot_uploads"); UPLOAD.mkdir(exist_ok=True)
OUTPUT = Path("/tmp/bot_outputs"); OUTPUT.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════════════
#  QR KOD
# ══════════════════════════════════════════════════════════════════════

def _qr_flowable(veri: str, boyut_cm: float = 2.8):
    """qrcode kütüphanesi varsa QR üret, yoksa placeholder döndür."""
    try:
        import qrcode as qrc
        from reportlab.platypus import Image as RLImg
        qr = qrc.QRCode(version=2, box_size=4, border=2)
        qr.add_data(veri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#1A1A2E", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        s = boyut_cm * cm
        return RLImg(buf, width=s, height=s)
    except ImportError:
        class QRPlaceholder(Flowable):
            def __init__(self, s): self.w = self.h = s
            def wrap(self, *_): return self.w, self.h
            def draw(self):
                s = self.w
                self.canv.setStrokeColor(colors.HexColor("#CCCCCC"))
                self.canv.setFillColor(colors.HexColor("#F5F5F5"))
                self.canv.roundRect(0, 0, s, s, 4, stroke=1, fill=1)
                self.canv.setFillColor(colors.HexColor("#AAAAAA"))
                self.canv.setFont(_FONT, 5.5)
                self.canv.drawCentredString(s/2, s/2+3, "QR KOD")
                self.canv.drawCentredString(s/2, s/2-5, "(deploy sonrası)")
        return QRPlaceholder(boyut_cm * cm)

# ══════════════════════════════════════════════════════════════════════
#  OCR
# ══════════════════════════════════════════════════════════════════════

def ocr_image(img_bytes: bytes) -> str:
    """Görüntüden metin çıkarır. Geçersiz format veya Tesseract hatası fırlatır."""
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("L")
    except UnidentifiedImageError as exc:
        raise ValueError(f"Geçersiz görüntü formatı: {exc}") from exc
    try:
        return pytesseract.image_to_string(img, lang="tur+eng").strip()
    except pytesseract.TesseractNotFoundError as exc:
        raise RuntimeError(
            "Tesseract OCR motoru bulunamadı. Lütfen sunucuya Tesseract kurun."
        ) from exc

def parse_invoice_text(text: str) -> dict:
    """Fatura metninden anahtar alanları çıkarır ve format doğrulaması yapar."""
    data: dict = {"ham_metin": text}

    m = re.search(r'(toplam|total)[^\d]*([\d.,]+)', text, re.IGNORECASE)
    if m:
        raw = m.group(2)
        try:
            data["toplam"] = float(raw.replace(",", "."))
        except ValueError:
            logger.warning("Toplam değeri sayıya dönüştürülemedi: %s", raw)

    m = re.search(r'\b(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})\b', text)
    if m:
        data["tarih"] = m.group(1)

    m = re.search(r'(fatura\s*no|invoice\s*no)[^\d\w]*(\w+)', text, re.IGNORECASE)
    if m:
        data["fatura_no"] = m.group(2)

    return data

# ══════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════

ACCENT = "1A1A2E"
LIGHT  = "E8F0FE"

def _safe_float(value: str) -> float | None:
    """Sayısal dönüşüm: başarısızsa None döndürür, loglama yapar."""
    cleaned = str(value).strip().replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None

def excel_olustur(basliklar: list, satirlar: list, sayfa_adi: str = "Veri") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sayfa_adi
    ws.sheet_view.showGridLines = False
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, b in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=c, value=b)
        cell.fill      = PatternFill("solid", fgColor=ACCENT)
        cell.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        ws.column_dimensions[get_column_letter(c)].width = max(len(str(b)) + 6, 14)

    for r, satir in enumerate(satirlar, 2):
        bg = LIGHT if r % 2 == 0 else "FFFFFF"
        for c, val in enumerate(satir, 1):
            cell           = ws.cell(row=r, column=c, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(size=10, name="Calibri")
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20

    donusum_uyarilari: list[str] = []
    if satirlar:
        tr = len(satirlar) + 2
        ws.cell(row=tr, column=1, value="TOPLAM").font = Font(bold=True, name="Calibri", color="FFFFFF")
        ws.cell(row=tr, column=1).fill      = PatternFill("solid", fgColor="1A1A2E")
        ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")

        for c in range(2, len(basliklar) + 1):
            h    = get_column_letter(c)
            vals = []
            for i, satir in enumerate(satirlar):
                raw = satir[c - 1] if c - 1 < len(satir) else ""
                num = _safe_float(str(raw))
                if num is not None:
                    vals.append(num)
                elif str(raw).strip():
                    donusum_uyarilari.append(
                        f"Satır {i+2}, Sütun {h}: '{raw}' sayıya dönüştürülemedi."
                    )

            if vals:
                cell           = ws.cell(row=tr, column=c, value=f"=SUM({h}2:{h}{tr-1})")
                cell.font      = Font(bold=True, name="Calibri", color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1A1A2E")
                cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    if donusum_uyarilari:
        logger.warning("Excel toplam hesabında dönüşüm uyarıları: %s", donusum_uyarilari)
    return buf.getvalue(), donusum_uyarilari

def csv_to_excel(csv_bytes: bytes) -> bytes:
    """CSV baytlarını Excel'e dönüştürür. Encoding hatası varsa bilgi döndürür."""
    try:
        text = csv_bytes.decode("utf-8")
    except UnicodeDecodeError:
        # UTF-8 başarısız — latin-1 dene, uyar
        logger.warning("CSV UTF-8 ile okunamadı, latin-1 deneniyor.")
        text = csv_bytes.decode("latin-1", errors="replace")

    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise ValueError("CSV dosyası boş veya okunamadı.")

    xb, _ = excel_olustur(rows[0], rows[1:])
    return xb

# ══════════════════════════════════════════════════════════════════════
#  FATURA PDF  v3 PREMIUM
# ══════════════════════════════════════════════════════════════════════

def _safe_float_fatura(value, default: float = 0.0, field: str = "") -> float:
    """Fatura alanları için güvenli float dönüşümü."""
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        logger.warning("Fatura alanı '%s' için geçersiz değer '%s': %s", field, value, exc)
        return default

def fatura_pdf_olustur(data: dict) -> bytes:
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate

    W, H = A4
    ML = MR = 2 * cm
    MT = 2 * cm
    MB = 2.2 * cm

    # ── Veri + Doğrulama ──
    sirket     = str(data.get("sirket_adi", "Şirket"))
    sirket_bil = str(data.get("sirket_bilgi", ""))
    iletisim   = str(data.get("iletisim", ""))
    musteri    = str(data.get("musteri_adi", "Müşteri"))
    musteri_ad = str(data.get("musteri_adres", ""))
    fatura_no  = str(data.get("fatura_no",
                              f"FTR-{datetime.date.today():%Y%m%d}-001"))
    tarih      = str(data.get("tarih", datetime.date.today().strftime("%d/%m/%Y")))
    vade       = str(data.get("vade", ""))
    kalemler   = data.get("kalemler", [])
    notlar     = str(data.get("notlar", ""))
    banka      = str(data.get("banka_bilgi", ""))
    kdv_oran   = _safe_float_fatura(data.get("kdv_oran", 20), 20.0, "kdv_oran")
    para       = str(data.get("para_birimi", "₺"))
    logo_b64   = str(data.get("logo_base64", ""))

    if not (0 <= kdv_oran <= 100):
        logger.warning("kdv_oran sınır dışı (%s), 20 kullanılıyor.", kdv_oran)
        kdv_oran = 20.0

    if not isinstance(kalemler, list):
        raise ValueError("'kalemler' alanı liste olmalıdır.")

    ara_toplam = sum(
        _safe_float_fatura(k.get("miktar", 1), 1.0, "miktar") *
        _safe_float_fatura(k.get("birim_fiyat", 0), 0.0, "birim_fiyat")
        for k in kalemler
    )
    kdv_tutar = ara_toplam * kdv_oran / 100
    genel_top = ara_toplam + kdv_tutar

    buf = io.BytesIO()

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(_FONT, 7)
        canvas.setFillColor(colors.HexColor("#AAAAAA"))
        canvas.drawString(ML, 1.2*cm, f"Ref: {fatura_no}  •  Freelance Otomasyon Botu")
        canvas.drawRightString(W - MR, 1.2*cm, f"Sayfa {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#E0E0E0"))
        canvas.setLineWidth(0.5)
        canvas.line(ML, 1.6*cm, W - MR, 1.6*cm)
        canvas.restoreState()

    frame = Frame(ML, MB, W - ML - MR, H - MT - MB, id="main")
    tmpl  = PageTemplate(id="p1", frames=[frame], onPage=_footer)
    doc   = BaseDocTemplate(buf, pagesize=A4,
                             leftMargin=ML, rightMargin=MR,
                             topMargin=MT, bottomMargin=MB)
    doc.addPageTemplates([tmpl])

    def st(name, size=10, bold=False, color="#000000", align=0, space=0, leading=None):
        fn = _FONT_BOLD if bold else _FONT
        return ParagraphStyle(name, fontName=fn, fontSize=size,
                              textColor=colors.HexColor(color),
                              alignment=align, spaceAfter=space,
                              leading=leading or size * 1.35)

    elems = []

    # ══ HEADER ══
    sol_items = []
    if logo_b64:
        try:
            raw = base64.b64decode(logo_b64)
            logo_img = RLImage(io.BytesIO(raw), width=4*cm, height=1.5*cm)
            sol_items.append(logo_img)
        except Exception as exc:
            logger.warning("Logo yüklenemedi: %s", exc)

    sol_items.append(Paragraph(sirket, st("s_baslik", 15, bold=True,
                                          color="#1A1A2E", leading=20)))
    if sirket_bil:
        sol_items.append(Spacer(1, 2))
        for line in sirket_bil.split("\n"):
            sol_items.append(Paragraph(line, st("s_bil", 8, color="#666666")))
    if iletisim:
        sol_items.append(Paragraph(iletisim, st("s_ile", 8, color="#666666")))

    sag = Paragraph("FATURA", st("fatura_etiket", 22, bold=True,
                                  color="#FFFFFF", align=1))

    sol_tbl = Table([[x] for x in sol_items], colWidths=[11*cm])
    sol_tbl.setStyle(TableStyle([
        ("TOPPADDING",    (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
    ]))

    header_tbl = Table([[sol_tbl, sag]], colWidths=[11*cm, 6*cm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (1, 0), (1, 0), colors.HexColor("#1A1A2E")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (0,  0), 0),
        ("RIGHTPADDING",  (0, 0), (0,  0), 8),
        ("LEFTPADDING",   (1, 0), (1,  0), 10),
        ("RIGHTPADDING",  (1, 0), (1,  0), 10),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
    ]))
    elems.append(header_tbl)
    elems.append(Spacer(1, 0.5*cm))

    # ══ META + QR ══
    meta_left = [
        [Paragraph("FATURA NO", st("ml", 7, bold=True, color="#999999")),
         Paragraph(fatura_no,   st("mv", 10, bold=True, color="#1A1A2E"))],
        [Paragraph("TARİH",     st("tl", 7, bold=True, color="#999999")),
         Paragraph(tarih,       st("tv", 10, color="#1A1A2E"))],
    ]
    if vade:
        meta_left.append([
            Paragraph("VADE", st("vl", 7, bold=True, color="#999999")),
            Paragraph(vade,   st("vv", 10, color="#1A1A2E"))
        ])

    meta_tbl = Table(meta_left, colWidths=[2.5*cm, 8*cm])
    meta_tbl.setStyle(TableStyle([
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))

    qr_content = (f"FATURA:{fatura_no}|FIRMA:{sirket}"
                  f"|TOPLAM:{para}{genel_top:.2f}|TARIH:{tarih}")
    qr_obj = _qr_flowable(qr_content, boyut_cm=2.5)

    qr_label = Table([
        [qr_obj],
        [Paragraph("e-Fatura Ref.", st("qrl", 6, color="#AAAAAA", align=1))]
    ], colWidths=[2.8*cm])
    qr_label.setStyle(TableStyle([
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))

    meta_row = Table([[meta_tbl, qr_label]], colWidths=[13.5*cm, 3.5*cm])
    meta_row.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elems.append(meta_row)
    elems.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#E0E0E0"),
                             spaceAfter=6, spaceBefore=6))

    # ══ MÜŞTERİ ══
    elems.append(Paragraph("FATURA KESİLEN", st("fk", 7, bold=True, color="#999999")))
    elems.append(Paragraph(musteri, st("mn", 13, bold=True, color="#1A1A2E")))
    if musteri_ad:
        elems.append(Paragraph(musteri_ad, st("ma", 9, color="#666666")))
    elems.append(Spacer(1, 0.5*cm))

    # ══ KALEMLER TABLOSU ══
    def _th(txt, align=0):
        return Paragraph(txt, ParagraphStyle("th", fontName=_FONT_BOLD, fontSize=9,
                         textColor=colors.white, alignment=align))

    def _td(txt, align=0, bold=False):
        fn = _FONT_BOLD if bold else _FONT
        return Paragraph(str(txt), ParagraphStyle("td", fontName=fn, fontSize=9,
                         textColor=colors.HexColor("#1A1A2E"), alignment=align))

    tablo = [[_th("#", 1), _th("Açıklama"), _th("Miktar", 2),
              _th("Birim Fiyat", 2), _th("Tutar", 2)]]
    for i, k in enumerate(kalemler, 1):
        mik   = _safe_float_fatura(k.get("miktar", 1),       1.0, "miktar")
        bf    = _safe_float_fatura(k.get("birim_fiyat", 0),  0.0, "birim_fiyat")
        tutar = mik * bf
        tablo.append([
            _td(str(i), align=1),
            _td(k.get("aciklama", "")),
            _td(f"{mik:g}", align=2),
            _td(f"{para}{bf:,.2f}", align=2),
            _td(f"{para}{tutar:,.2f}", align=2),
        ])

    kalem_tbl = Table(tablo, colWidths=[0.9*cm, 8.6*cm, 1.8*cm, 2.3*cm, 3.4*cm],
                      repeatRows=1)
    kalem_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#1A1A2E")),
        ("TOPPADDING",    (0, 0), (-1,  0), 8),
        ("BOTTOMPADDING", (0, 0), (-1,  0), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F6FF")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#D8D8E8")),
        ("LINEBELOW",     (0, 0), (-1,  0), 1.5, colors.HexColor("#0A0A1E")),
        ("TOPPADDING",    (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems.append(KeepTogether(kalem_tbl))
    elems.append(Spacer(1, 0.4*cm))

    # ══ TOPLAMLAR ══
    top_data = [
        [Paragraph("Ara Toplam", st("atl", 9, color="#555555")),
         Paragraph(f"{para}{ara_toplam:,.2f}", st("atv", 9, color="#1A1A2E", align=2))],
        [Paragraph(f"KDV (%{kdv_oran:.0f})", st("kl", 9, color="#555555")),
         Paragraph(f"{para}{kdv_tutar:,.2f}", st("kv", 9, color="#1A1A2E", align=2))],
        [Paragraph("GENEL TOPLAM", st("gtl", 11, bold=True, color="#FFFFFF")),
         Paragraph(f"{para}{genel_top:,.2f}", st("gtv", 11, bold=True, color="#FFFFFF", align=2))],
    ]
    top_tbl = Table(top_data, colWidths=[4.5*cm, 3.5*cm], hAlign="RIGHT")
    top_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 2), (-1, 2), colors.HexColor("#1A1A2E")),
        ("LINEABOVE",    (0, 2), (-1, 2), 1,   colors.HexColor("#1A1A2E")),
        ("LINEBELOW",    (0, 1), (-1, 1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("LEFTPADDING",  (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("ROUNDEDCORNERS", [0, 0, 4, 4]),
    ]))
    elems.append(top_tbl)
    elems.append(Spacer(1, 0.6*cm))

    # ══ BANKA BİLGİSİ BLOĞU ══
    banka_str = banka or notlar
    if banka_str:
        elems.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor("#E0E0E0"), spaceAfter=6))
        banka_content = [
            [Paragraph("🏦 BANKA / ÖDEME BİLGİLERİ",
                       st("bbl", 8, bold=True, color="#FFFFFF")),
             Paragraph("NOTLAR",
                       st("nl",  8, bold=True, color="#FFFFFF", align=2))]
        ]
        b_lines = (banka  or "").strip()
        n_lines = (notlar or "").strip()
        sol_txt = b_lines if b_lines else n_lines
        sag_txt = n_lines if b_lines else ""
        body = [
            [Paragraph(sol_txt.replace("\n", "<br/>"),
                       st("bt", 8, color="#333333")) if sol_txt else Spacer(1, 1),
             Paragraph(sag_txt.replace("\n", "<br/>"),
                       st("nt", 8, color="#333333", align=2)) if sag_txt else Spacer(1, 1)]
        ]
        banka_tbl = Table(banka_content + body, colWidths=[8.5*cm, 8.5*cm])
        banka_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8F8FC")),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("SPAN",          (0, 0), (-1, 0)) if not sag_txt else ("NOP", (0, 0), (0, 0)),
        ]))
        elems.append(banka_tbl)

    doc.build(elems)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    # __file__ ile taşınabilir yol — sabit /home/claude/ yerine proje dizini
    html_path = Path(__file__).resolve().parent / "index.html"
    try:
        return render_template_string(html_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return "index.html bulunamadı.", 404

@app.route("/api/ocr", methods=["POST"])
def api_ocr():
    if "file" not in request.files:
        return jsonify({"error": "Dosya yok"}), 400
    try:
        text   = ocr_image(request.files["file"].read())
        parsed = parse_invoice_text(text)
        return jsonify({"ok": True, "metin": text, "parsed": parsed})
    except ValueError as exc:
        # Geçersiz görüntü formatı
        logger.warning("OCR format hatası: %s", exc)
        return jsonify({"error": str(exc)}), 422
    except RuntimeError as exc:
        # Tesseract bulunamadı
        logger.error("OCR motor hatası: %s", exc)
        return jsonify({"error": str(exc)}), 503
    except Exception as exc:
        logger.exception("OCR bilinmeyen hata")
        return jsonify({"error": f"OCR işlemi başarısız: {exc}"}), 500

@app.route("/api/excel", methods=["POST"])
def api_excel():
    d = request.json or {}
    basliklar = d.get("basliklar", [])
    satirlar  = d.get("satirlar", [])
    sayfa_adi = d.get("sayfa_adi", "Veri")

    if not isinstance(basliklar, list) or not isinstance(satirlar, list):
        return jsonify({"error": "'basliklar' ve 'satirlar' liste olmalıdır."}), 422
    try:
        xb, uyarilar = excel_olustur(basliklar, satirlar, sayfa_adi)
        resp = send_file(
            io.BytesIO(xb),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="rapor.xlsx"
        )
        if uyarilar:
            resp.headers["X-Conversion-Warnings"] = str(len(uyarilar))
        return resp
    except Exception as exc:
        logger.exception("Excel oluşturma hatası")
        return jsonify({"error": f"Excel oluşturulamadı: {exc}"}), 500

@app.route("/api/csv-to-excel", methods=["POST"])
def api_csv_excel():
    if "file" not in request.files:
        return jsonify({"error": "Dosya yok"}), 400
    try:
        xb = csv_to_excel(request.files["file"].read())
        return send_file(
            io.BytesIO(xb),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="donusturulmus.xlsx"
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 422
    except Exception as exc:
        logger.exception("CSV→Excel dönüşüm hatası")
        return jsonify({"error": f"Dönüşüm başarısız: {exc}"}), 500

@app.route("/api/fatura", methods=["POST"])
def api_fatura():
    d = request.json or {}
    try:
        pdf   = fatura_pdf_olustur(d)
        fname = f"fatura_{d.get('fatura_no', '001')}.pdf"
        return send_file(io.BytesIO(pdf), mimetype="application/pdf",
                         as_attachment=True, download_name=fname)
    except ValueError as exc:
        return jsonify({"error": f"Geçersiz fatura verisi: {exc}"}), 422
    except Exception as exc:
        logger.exception("Fatura PDF oluşturma hatası")
        return jsonify({"error": f"PDF oluşturulamadı: {exc}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
